import openpyxl
import os


delimiter = os.sep

def getDataFromExcel(excelFileName: str) -> dict: # {"Temp" : [1.0,2.0], "Pressure" : [9.0, 8.0]}
    workbook = None
    try:
        # excelFileName = "test.xlsx"
        workbook = openpyxl.load_workbook(
            excelFileName)  # очень важно!! в названии листа не должно быть нижнего подчеркивнаия "_" иначе программа ломается
        values = readData(workbook)
        return values
    except ValueError as ex:
        print("ОШИБКА")
        print(ex)
    finally:
        if workbook is not None:
            workbook.close()


def saveParamsToExcel(excelFileName: str, allOutputValues: dict): #!! ALL i.e. {"Температура":[2.0 9.0], "Давление":[2.0 9.0]}
    workbook = None
    try:
        workbook = openpyxl.load_workbook(excelFileName)
        sheet = workbook[workbook.sheetnames[1]] # second list

        # set params names to first row in Excel
        colIdx = 1
        for paramName in allOutputValues:
            sheet.cell(row=1, column=colIdx).value = paramName
            colIdx += 1

        #set all params
        colIdx = 1
        for paramName in allOutputValues:
            vals = allOutputValues[paramName] # it's a list
            rowIdx = 2
            for i in range (len(vals)):
                sheet.cell(row=rowIdx, column=colIdx).value = vals[i]
                rowIdx += 1
            colIdx += 1
        workbook.save(excelFileName)
    except PermissionError as e:
        print("!!!!ОШИБКА!!!!")
        print("Ошибка доступа: Файл " + excelFileName + " открыт (закройте его)")
    except Exception as ex:
        print("!!!!ОШИБКА!!!!")
        print(ex)
    finally:
        if workbook is not None:
            workbook.close()


def readData(workbook) -> dict:
    paramNames = []
    data = {}
    sheet = workbook[workbook.sheetnames[0]]
    # process the column names
    for colI in range(1, sheet.max_column + 1):
        colName = sheet.cell(row=1, column=colI).value
        paramNames.append(colName)

    for rowI in range(2, sheet.max_row + 1):
        for colI in range(1, sheet.max_column + 1):
            val = float(str(sheet.cell(row=rowI, column=colI).value).replace(",", "."))
            # data[paramNames[colI-1]] = val
            try:
                data[paramNames[colI - 1]].append(val)
            except KeyError:
                data[paramNames[colI - 1]] = [val]
    return data

