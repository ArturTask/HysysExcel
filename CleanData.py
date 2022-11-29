import datetime
import os
import traceback

import openpyxl
from config.conf import Paths
import pandas as pd
from os import listdir
from os.path import isfile, join

delimiter = os.sep


excelFolder = Paths.EXCEL_FOLDER.value


def getExcelFilesFromFolder(folderPath: str):
    onlyfiles = [f for f in listdir(folderPath) if
                 (isfile(join(folderPath, f)) and f.endswith(".xlsx") and not f.startswith("~$"))]
    return onlyfiles


def getExcelNameFromFolder() -> str:
    openedExcelFiles = getExcelFilesFromFolder(excelFolder)
    if len(openedExcelFiles) == 1:
        return openedExcelFiles[0]
    else:
        raise ValueError('Более 1 .xlsx файла (или ни одного) в папке ' + excelFolder)


def clean(workbook):
    deleteRows = []
    sheet = workbook[workbook.sheetnames[0]]
    for i in range(6, sheet.max_row + 1):
        # define emptiness of cell
        val = float(sheet.cell(row=i, column=2).value.replace(",", "."))
        print("ряд " + str(i) + " время: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S") + "\tзначение = " + str(val))
        if val < 5.8 or val > 5.9:
            # collect indexes of rows
            deleteRows.append(i)

    cleanDeleteRows = cleanRows(deleteRows)
    for key in cleanDeleteRows:
        print("Удаляем " + str(key) + "   " + str(cleanDeleteRows[key]) + " время: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
        amount = cleanDeleteRows[key]
        sheet.delete_rows(idx=key, amount=amount)


def cleanRows(deleteRows) ->dict : # we need to get dict with REVERSE order (from end to beginning in order to delete properly and not mix indexes after deleting rows)
    result = {} # {index: amount of rows to delete}
    amount = 1

    for i in range(len(deleteRows)-2,-1,-1):
        prevRowIdx = deleteRows[i]
        rowIdx = deleteRows[i+1]
        if (rowIdx -1 == prevRowIdx):
            amount += 1
        else:
            result[rowIdx] = amount
            amount = 1

    result[deleteRows[0]] = amount

    print(deleteRows)
    print(result)

    return result

workbook = None
try:
    currentExcelFile = getExcelNameFromFolder()
    print("Смотрим этот файл: " + currentExcelFile)
    print("начало - " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    workbook = openpyxl.load_workbook(currentExcelFile)  # очень важно!! в названии листа не должно быть нижнего подчеркивнаия "_" иначе программа ломается
    print("workbook - " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    clean(workbook)
    workbook.save(excelFolder + delimiter + currentExcelFile.replace(".xlsx", "_" + datetime.datetime.now().strftime("%d-%m-%Y(%H;%M)") + ".xlsx"))
except ValueError as ex:
    print("ОШИБКА")
    print(ex)
finally:
    if workbook is not None:
        workbook.close()
